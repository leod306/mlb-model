"""
daily_picks_tracker.py
-----------------------
Single script that does three jobs in order:

  1. EVALUATE yesterday's picks against actual results
  2. SAVE today's predictions as picks to evaluate tomorrow
  3. EXPORT full picks history to Excel with color coding + charts

Excel file saved to: reports/picks_tracker.xlsx
"""
from __future__ import annotations

import os
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if os.getenv("DYNO") is None:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(PROJECT_ROOT, ".env"), override=False)

DATABASE_URL = (
    os.getenv("DATABASE_URL", "")
    .replace("postgres://", "postgresql://", 1)
    .replace("postgresql+psycopg2://", "postgresql://", 1)
)
GAMES_TABLE = os.getenv("MLB_GAMES_TABLE", "games")
PRED_TABLE  = os.getenv("MLB_PREDICTIONS_TABLE", "predictions")
PICKS_TABLE = "daily_picks"

DEFAULT_OU_LINE = 8.5

REPORTS_DIR = Path(PROJECT_ROOT) / "reports"
EXCEL_PATH  = REPORTS_DIR / "picks_tracker.xlsx"

# Colors
GREEN_FILL  = "C6EFCE"; GREEN_FONT  = "276221"
RED_FILL    = "FFC7CE"; RED_FONT    = "9C0006"
YELLOW_FILL = "FFEB9C"; YELLOW_FONT = "9C5700"
BLUE_FILL   = "DEEAF1"; BLUE_FONT   = "1F4E79"
GRAY_FILL   = "F2F2F2"; DARK_HEADER = "1F4E79"
WHITE       = "FFFFFF"; LIGHT_GREEN = "EAF3DE"
LIGHT_RED   = "FCEBEB"; LIGHT_GRAY  = "F9F9F9"
NAVY        = "1B2A4A"; STEEL       = "4A6FA5"
MID         = "2C3E50"; ACCENT      = "F0A500"


# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------

def conn():
    return psycopg2.connect(DATABASE_URL)


def ensure_tables(cur):
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {PICKS_TABLE} (
            id                SERIAL PRIMARY KEY,
            game_pk           BIGINT NOT NULL,
            pick_date         DATE   NOT NULL,
            home_team         TEXT,
            away_team         TEXT,
            home_sp           TEXT,
            away_sp           TEXT,
            ml_pick           TEXT,
            runline_pick      TEXT,
            ou_pick           TEXT,
            home_win_prob     DOUBLE PRECISION,
            away_win_prob     DOUBLE PRECISION,
            pred_run_diff     DOUBLE PRECISION,
            pred_total_runs   DOUBLE PRECISION,
            market_total_line DOUBLE PRECISION,
            home_ml_implied   INT,
            away_ml_implied   INT,
            home_score        INT,
            away_score        INT,
            actual_run_diff   DOUBLE PRECISION,
            actual_total      DOUBLE PRECISION,
            ml_correct        BOOLEAN,
            runline_correct   BOOLEAN,
            ou_correct        BOOLEAN,
            evaluated         BOOLEAN DEFAULT FALSE,
            evaluated_at      TIMESTAMPTZ,
            created_at        TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            UNIQUE (game_pk, pick_date)
        );
    """)
    cur.execute(f"ALTER TABLE {PICKS_TABLE} ADD COLUMN IF NOT EXISTS market_total_line DOUBLE PRECISION;")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_picks_date ON {PICKS_TABLE}(pick_date);")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_picks_eval ON {PICKS_TABLE}(evaluated);")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sym(val):
    if val is None: return "—"
    return "✅" if val else "❌"


def pct_str(wins, total):
    if not total: return "—"
    return f"{int(wins)}/{total} ({round(wins / total * 100, 1)}%)"


def result_fill(val):
    if val is True:  return PatternFill("solid", fgColor=GREEN_FILL)
    if val is False: return PatternFill("solid", fgColor=RED_FILL)
    return PatternFill("solid", fgColor=YELLOW_FILL)


def result_font(val):
    if val is True:  return Font(name="Arial", size=10, color=GREEN_FONT, bold=True)
    if val is False: return Font(name="Arial", size=10, color=RED_FONT, bold=True)
    return Font(name="Arial", size=10, color=YELLOW_FONT)


def header_cell(cell, text, size=10):
    cell.value     = text
    cell.font      = Font(bold=True, color=WHITE, name="Arial", size=size)
    cell.fill      = PatternFill("solid", fgColor=DARK_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def win_prob_fill(prob):
    if prob is None:   return PatternFill("solid", fgColor=GRAY_FILL)
    if prob >= 0.60:   return PatternFill("solid", fgColor=GREEN_FILL)
    if prob >= 0.50:   return PatternFill("solid", fgColor=YELLOW_FILL)
    return PatternFill("solid", fgColor=RED_FILL)


def win_prob_font(prob):
    if prob is None:   return Font(name="Arial", size=10)
    if prob >= 0.60:   return Font(name="Arial", size=10, color=GREEN_FONT, bold=True)
    if prob >= 0.50:   return Font(name="Arial", size=10, color=YELLOW_FONT, bold=True)
    return Font(name="Arial", size=10, color=RED_FONT, bold=True)


def pct_fill(pct_val):
    if pct_val >= 57:    return PatternFill("solid", fgColor=GREEN_FILL)
    if pct_val >= 52.4:  return PatternFill("solid", fgColor=YELLOW_FILL)
    return PatternFill("solid", fgColor=RED_FILL)


def pct_font(pct_val):
    if pct_val >= 57:    return Font(name="Arial", size=10, color=GREEN_FONT, bold=True)
    if pct_val >= 52.4:  return Font(name="Arial", size=10, color=YELLOW_FONT, bold=True)
    return Font(name="Arial", size=10, color=RED_FONT, bold=True)


def hdr(ws, cell_ref, val, bg=NAVY, fg=WHITE, bold=True, size=10, wrap=False, center=True):
    c = ws[cell_ref]
    c.value = val
    c.font  = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill  = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap
    )


def scell(ws, ref, val, bold=False, bg=None, fg="000000", size=10, fmt=None, center=False, italic=False):
    c = ws[ref]
    c.value = val
    c.font  = Font(name="Arial", bold=bold, color=fg, size=size, italic=italic)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    if fmt:
        c.number_format = fmt


def thin_border_block(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="BDC3C7")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# Evaluation logic
# ---------------------------------------------------------------------------

def eval_ml(ml_pick, home_team, away_team, home_score, away_score):
    if None in (home_score, away_score):        return None
    if not ml_pick or ml_pick in ("PASS", ""):  return None
    if home_score == away_score:                 return None
    winner = home_team if home_score > away_score else away_team
    return ml_pick == winner


def eval_runline(runline_pick, home_team, away_team, home_score, away_score):
    if None in (home_score, away_score):                  return None
    if not runline_pick or runline_pick in ("PASS", ""):  return None
    try:
        parts = str(runline_pick).rsplit(" ", 1)
        if len(parts) != 2: return None
        team, spread_str = parts
        spread = float(spread_str)
    except Exception:
        return None
    diff = home_score - away_score
    if team == home_team:
        adjusted = diff + spread
        return None if adjusted == 0 else adjusted > 0
    if team == away_team:
        adjusted = (-diff) + spread
        return None if adjusted == 0 else adjusted > 0
    return None


def eval_ou(ou_pick, home_score, away_score, market_line, pred_total):
    if None in (home_score, away_score):          return None
    if not ou_pick or ou_pick in ("PASS", ""):    return None
    actual = home_score + away_score
    if market_line and market_line > 0:
        line = market_line
    elif pred_total and pred_total > 0:
        line = pred_total
    else:
        line = DEFAULT_OU_LINE
    if actual == line: return None
    return (ou_pick == "OVER" and actual > line) or (ou_pick == "UNDER" and actual < line)


def predicted_scores(home_win_prob, pred_run_diff, pred_total_runs):
    if pred_total_runs is None or pred_run_diff is None:
        return None, None
    home = (pred_total_runs + pred_run_diff) / 2
    away = (pred_total_runs - pred_run_diff) / 2
    return round(home, 1), round(away, 1)


# ---------------------------------------------------------------------------
# Step 1: Evaluate yesterday
# ---------------------------------------------------------------------------

def evaluate_yesterday(cur):
    yesterday = date.today() - timedelta(days=1)
    print(f"{'='*55}")
    print(f"  STEP 1 — Evaluating picks for {yesterday}")
    print(f"{'='*55}\n")

    cur.execute(f"""
        SELECT id, game_pk, home_team, away_team,
               ml_pick, runline_pick, ou_pick,
               pred_total_runs, market_total_line
        FROM {PICKS_TABLE}
        WHERE pick_date = %s AND evaluated = FALSE
        ORDER BY game_pk
    """, (yesterday,))
    picks = cur.fetchall()

    if not picks:
        print(f"  No unevaluated picks for {yesterday}.\n")
        return

    ml_res, rl_res, ou_res = [], [], []

    for pick in picks:
        pick_id, game_pk, home_team, away_team, \
        ml_pick, runline_pick, ou_pick, pred_total, market_line = pick

        cur.execute(f"SELECT home_score, away_score FROM {GAMES_TABLE} WHERE game_pk = %s", (game_pk,))
        result = cur.fetchone()

        if not result or result[0] is None or result[1] is None:
            print(f"  {away_team} @ {home_team}: no result yet, skipping")
            continue

        home_score, away_score = result
        actual_total    = home_score + away_score
        actual_run_diff = home_score - away_score

        ml_correct = eval_ml(ml_pick, home_team, away_team, home_score, away_score)
        rl_correct = eval_runline(runline_pick, home_team, away_team, home_score, away_score)
        ou_correct = eval_ou(ou_pick, home_score, away_score, market_line, pred_total)

        if ou_pick and ou_pick not in ("PASS", ""):
            ou_line_used = market_line if (market_line and market_line > 0) else (pred_total or DEFAULT_OU_LINE)
            ou_line_src  = "Vegas" if (market_line and market_line > 0) else "Model"
            ou_display   = f"({ou_line_src} line: {ou_line_used} | actual: {actual_total})"
        else:
            ou_display = "(PASS — not graded)"

        print(f"  {away_team} ({away_score}) @ {home_team} ({home_score})")
        print(f"    ML:  {(ml_pick or '—'):<20} {sym(ml_correct)}")
        print(f"    RL:  {(runline_pick or '—'):<20} {sym(rl_correct)}")
        print(f"    O/U: {(ou_pick or '—'):<20} {sym(ou_correct)}  {ou_display}\n")

        if ml_correct is not None: ml_res.append(ml_correct)
        if rl_correct is not None: rl_res.append(rl_correct)
        if ou_correct is not None: ou_res.append(ou_correct)

        cur.execute(f"""
            UPDATE {PICKS_TABLE} SET
                home_score=%(hs)s, away_score=%(as_)s,
                actual_run_diff=%(ard)s, actual_total=%(at_)s,
                ml_correct=%(ml)s, runline_correct=%(rl)s, ou_correct=%(ou)s,
                evaluated=TRUE, evaluated_at=NOW()
            WHERE id=%(id)s
        """, {
            "hs": home_score, "as_": away_score,
            "ard": actual_run_diff, "at_": actual_total,
            "ml": ml_correct, "rl": rl_correct, "ou": ou_correct,
            "id": pick_id,
        })

    print(f"  --- {yesterday} Summary ---")
    print(f"  Moneyline : {pct_str(sum(ml_res), len(ml_res))}")
    print(f"  Run Line  : {pct_str(sum(rl_res), len(rl_res))}")
    print(f"  Over/Under: {pct_str(sum(ou_res), len(ou_res))}\n")

    cur.execute(f"""
        SELECT
            COUNT(*) FILTER (WHERE ml_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN ml_correct THEN 1 ELSE 0 END), 0),
            COUNT(*) FILTER (WHERE runline_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN runline_correct THEN 1 ELSE 0 END), 0),
            COUNT(*) FILTER (WHERE ou_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN ou_correct THEN 1 ELSE 0 END), 0)
        FROM {PICKS_TABLE} WHERE evaluated = TRUE
    """)
    row = cur.fetchone()
    if row:
        ml_tot, ml_w, rl_tot, rl_w, ou_tot, ou_w = row
        print("  --- Season Totals ---")
        print(f"  Moneyline : {pct_str(ml_w, ml_tot)}")
        print(f"  Run Line  : {pct_str(rl_w, rl_tot)}")
        print(f"  Over/Under: {pct_str(ou_w, ou_tot)}")
        print("  Break-even: 52.4% | Sharp target: 57%+\n")


# ---------------------------------------------------------------------------
# Step 2: Save today's picks
# ---------------------------------------------------------------------------

def save_today_picks(cur):
    today = date.today()
    print(f"{'='*55}")
    print(f"  STEP 2 — Saving picks for {today}")
    print(f"{'='*55}\n")

    cur.execute(f"""
        SELECT
            g.game_pk, g.home_team, g.away_team,
            COALESCE(gp.home_sp_name, g.home_starting_pitcher),
            COALESCE(gp.away_sp_name, g.away_starting_pitcher),
            p.ml_pick, p.runline_pick, p.ou_pick,
            p.home_win_prob, p.away_win_prob,
            p.run_diff_pred, p.total_runs_pred,
            p.market_total_line,
            p.home_ml_implied, p.away_ml_implied
        FROM {GAMES_TABLE} g
        LEFT JOIN {PRED_TABLE} p ON p.game_pk = g.game_pk
        LEFT JOIN game_probables gp ON gp.game_pk = g.game_pk
        WHERE g.official_date = %s
          AND g.game_type = 'R'
          AND p.ml_pick IS NOT NULL
        ORDER BY g.game_pk
    """, (today,))
    rows = cur.fetchall()

    if not rows:
        print(f"  No predictions for {today}.\n")
        return

    data = []
    for r in rows:
        (game_pk, home_team, away_team, home_sp, away_sp,
         ml_pick, runline_pick, ou_pick,
         home_win_prob, away_win_prob,
         run_diff_pred, total_runs_pred, market_total_line,
         home_ml_implied, away_ml_implied) = r

        ou_line_str = f"Vegas: {market_total_line}" if market_total_line else f"Model: {round(total_runs_pred,1) if total_runs_pred else '—'}"
        print(f"  {away_team} @ {home_team}  |  ML: {ml_pick} | RL: {runline_pick} | O/U: {ou_pick} ({ou_line_str})")

        data.append((
            int(game_pk), today,
            home_team, away_team, home_sp, away_sp,
            ml_pick, runline_pick, ou_pick,
            home_win_prob, away_win_prob,
            run_diff_pred, total_runs_pred, market_total_line,
            home_ml_implied, away_ml_implied,
        ))

    execute_values(cur, f"""
        INSERT INTO {PICKS_TABLE} (
            game_pk, pick_date, home_team, away_team, home_sp, away_sp,
            ml_pick, runline_pick, ou_pick,
            home_win_prob, away_win_prob,
            pred_run_diff, pred_total_runs, market_total_line,
            home_ml_implied, away_ml_implied
        ) VALUES %s
        ON CONFLICT (game_pk, pick_date) DO NOTHING;
    """, data)
    # DO NOTHING: first write wins. Re-running the tracker later in the day
    # (after official lineups / odds update) will NOT overwrite the morning picks.
    # To reset a day: DELETE FROM daily_picks WHERE pick_date = 'YYYY-MM-DD';

    print(f"\n  ✅ Saved {len(data)} picks for {today}\n")


# ---------------------------------------------------------------------------
# Step 3: Export to Excel  (All Picks + Summary + Charts & Trends)
# ---------------------------------------------------------------------------

def export_excel(cur):
    print(f"{'='*55}")
    print(f"  STEP 3 — Exporting to Excel")
    print(f"{'='*55}\n")

    cur.execute(f"""
        SELECT
            pick_date, away_team, home_team, away_sp, home_sp,
            ml_pick, home_win_prob, away_win_prob,
            runline_pick, pred_run_diff,
            ou_pick, pred_total_runs, market_total_line,
            home_score, away_score,
            actual_total, actual_run_diff,
            ml_correct, runline_correct, ou_correct,
            evaluated
        FROM {PICKS_TABLE}
        ORDER BY pick_date DESC, game_pk
    """)
    rows = cur.fetchall()

    if not rows:
        print("  No picks to export yet.\n")
        return

    REPORTS_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: All Picks ──────────────────────────────────────────
    ws = wb.active
    ws.title = "All Picks"

    headers = [
        "Date", "Matchup", "Away SP", "Home SP",
        "ML Pick", "Win Prob", "Pred Score", "Actual Score",
        "RL Pick", "Run Diff",
        "O/U Pick", "Model Total", "Vegas Line",
        "ML ✓", "RL ✓", "O/U ✓",
    ]
    for col, h in enumerate(headers, 1):
        header_cell(ws.cell(row=1, column=col), h)
    ws.row_dimensions[1].height = 32

    for row_idx, r in enumerate(rows, 2):
        (pick_date, away_team, home_team, away_sp, home_sp,
         ml_pick, home_win_prob, away_win_prob,
         runline_pick, pred_run_diff,
         ou_pick, pred_total_runs, market_total_line,
         home_score, away_score,
         actual_total, actual_run_diff,
         ml_correct, runline_correct, ou_correct,
         evaluated) = r

        pred_home, pred_away = predicted_scores(home_win_prob, pred_run_diff, pred_total_runs)
        pred_score_str  = f"{pred_away}-{pred_home}" if pred_home is not None else "—"
        actual_score_str = f"{away_score}-{home_score}" if home_score is not None and away_score is not None else "—"
        matchup  = f"{away_team} @ {home_team}"
        win_prob = f"{round(home_win_prob * 100, 1)}%" if home_win_prob is not None else "—"

        if evaluated and ml_correct is True:    row_bg = LIGHT_GREEN
        elif evaluated and ml_correct is False: row_bg = LIGHT_RED
        else: row_bg = WHITE if row_idx % 2 == 0 else LIGHT_GRAY

        values = [
            str(pick_date), matchup, away_sp or "TBD", home_sp or "TBD",
            ml_pick or "—", win_prob, pred_score_str, actual_score_str,
            runline_pick or "—",
            round(pred_run_diff, 2)     if pred_run_diff     is not None else "—",
            ou_pick or "—",
            round(pred_total_runs, 1)   if pred_total_runs   is not None else "—",
            round(market_total_line, 1) if market_total_line is not None else "—",
            "✅" if ml_correct      else ("❌" if ml_correct      is False else "—"),
            "✅" if runline_correct  else ("❌" if runline_correct  is False else "—"),
            "✅" if ou_correct       else ("❌" if ou_correct       is False else "—"),
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.font      = Font(name="Arial", size=10)
            if col == 6:
                c.fill = win_prob_fill(home_win_prob)
                c.font = win_prob_font(home_win_prob)
            elif col == 14:
                c.fill = result_fill(ml_correct) if evaluated else PatternFill("solid", fgColor=GRAY_FILL)
                c.font = result_font(ml_correct)  if evaluated else Font(name="Arial", size=10)
            elif col == 15:
                c.fill = result_fill(runline_correct) if evaluated else PatternFill("solid", fgColor=GRAY_FILL)
                c.font = result_font(runline_correct)  if evaluated else Font(name="Arial", size=10)
            elif col == 16:
                c.fill = result_fill(ou_correct) if evaluated else PatternFill("solid", fgColor=GRAY_FILL)
                c.font = result_font(ou_correct)  if evaluated else Font(name="Arial", size=10)
            elif col == 13:
                c.fill = PatternFill("solid", fgColor=BLUE_FILL)
                c.font = Font(name="Arial", size=10, color=BLUE_FONT)

    widths = [12, 20, 18, 18, 12, 10, 12, 12, 14, 10, 10, 12, 11, 8, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"

    # ── Sheet 2: Summary ────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    ws2["B2"] = "MLB Model — Season Summary"
    ws2["B2"].font      = Font(bold=True, size=16, name="Arial", color=DARK_HEADER)
    ws2["B2"].alignment = Alignment(vertical="center")
    ws2.row_dimensions[2].height = 30
    ws2.merge_cells("B2:H2")

    season_headers = ["Metric", "Wins", "Total", "Win %", "Edge", "Status"]
    for col, h in enumerate(season_headers, 2):
        header_cell(ws2.cell(row=4, column=col), h)
    ws2.row_dimensions[4].height = 24

    cur.execute(f"""
        SELECT
            COUNT(*) FILTER (WHERE ml_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN ml_correct THEN 1 ELSE 0 END), 0),
            COUNT(*) FILTER (WHERE runline_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN runline_correct THEN 1 ELSE 0 END), 0),
            COUNT(*) FILTER (WHERE ou_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN ou_correct THEN 1 ELSE 0 END), 0)
        FROM {PICKS_TABLE} WHERE evaluated = TRUE
    """)
    s = cur.fetchone()

    if s:
        ml_tot, ml_w, rl_tot, rl_w, ou_tot, ou_w = s
        metrics = [
            ("Moneyline",  int(ml_w), int(ml_tot)),
            ("Run Line",   int(rl_w), int(rl_tot)),
            ("Over/Under", int(ou_w), int(ou_tot)),
        ]
        for row_i, (label, wins, total) in enumerate(metrics, 5):
            pct_val  = round(wins / total * 100, 1) if total else 0
            edge_val = round(pct_val - 52.4, 1)

            cells_vals = [label, wins, total, f"{pct_val}%", f"{'+' if edge_val >= 0 else ''}{edge_val}%",
                          "Above break-even ✓" if edge_val >= 0 else "Below break-even ✗"]
            for col_i, val in enumerate(cells_vals, 2):
                c = ws2.cell(row=row_i, column=col_i, value=val)
                c.font      = Font(name="Arial", size=11, bold=(col_i == 2))
                c.alignment = Alignment(horizontal="center" if col_i > 2 else "left", vertical="center")
                c.border    = Border(
                    left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                    top=Side(style="thin", color="DDDDDD"),  bottom=Side(style="thin", color="DDDDDD"),
                )
                if col_i == 5:
                    c.fill = pct_fill(pct_val); c.font = pct_font(pct_val)
                elif col_i == 6:
                    c.fill = PatternFill("solid", fgColor=GREEN_FILL if edge_val >= 0 else RED_FILL)
                    c.font = Font(name="Arial", size=11, bold=True, color=GREEN_FONT if edge_val >= 0 else RED_FONT)
                elif col_i == 7:
                    c.fill = PatternFill("solid", fgColor=GREEN_FILL if edge_val >= 0 else RED_FILL)
                    c.font = Font(name="Arial", size=10, color=GREEN_FONT if edge_val >= 0 else RED_FONT)
                else:
                    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            ws2.row_dimensions[row_i].height = 22

    note_cell = ws2.cell(row=9, column=2, value="Break-even threshold: 52.4% | Sharp target: 57%+")
    note_cell.font      = Font(name="Arial", size=10, italic=True, color="888888")
    note_cell.alignment = Alignment(horizontal="left")
    ws2.merge_cells("B9:H9")

    ws2.cell(row=11, column=2, value="Daily Breakdown").font = Font(bold=True, size=13, name="Arial", color=DARK_HEADER)
    ws2.row_dimensions[11].height = 24

    daily_headers = ["Date", "ML W", "ML Tot", "ML %", "RL W", "RL Tot", "RL %", "O/U W", "O/U Tot", "O/U %"]
    for col, h in enumerate(daily_headers, 2):
        header_cell(ws2.cell(row=12, column=col), h)
    ws2.row_dimensions[12].height = 22

    cur.execute(f"""
        SELECT pick_date,
            COUNT(*) FILTER (WHERE ml_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN ml_correct THEN 1 ELSE 0 END), 0),
            COUNT(*) FILTER (WHERE runline_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN runline_correct THEN 1 ELSE 0 END), 0),
            COUNT(*) FILTER (WHERE ou_correct IS NOT NULL),
            COALESCE(SUM(CASE WHEN ou_correct THEN 1 ELSE 0 END), 0)
        FROM {PICKS_TABLE}
        WHERE evaluated = TRUE
        GROUP BY pick_date
        ORDER BY pick_date DESC
    """)
    daily = cur.fetchall()

    for row_idx, d in enumerate(daily, 13):
        pick_date, ml_tot, ml_w, rl_tot, rl_w, ou_tot, ou_w = d
        mp = round(ml_w / ml_tot * 100, 1) if ml_tot else 0
        rp = round(rl_w / rl_tot * 100, 1) if rl_tot else 0
        op = round(ou_w / ou_tot * 100, 1) if ou_tot else 0

        vals = [str(pick_date), int(ml_w), int(ml_tot), f"{mp}%" if ml_tot else "—",
                int(rl_w), int(rl_tot), f"{rp}%" if rl_tot else "—",
                int(ou_w), int(ou_tot), f"{op}%" if ou_tot else "—"]
        bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
        for col_i, val in enumerate(vals, 2):
            c = ws2.cell(row=row_idx, column=col_i, value=val)
            c.font      = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if col_i > 2 else "left", vertical="center")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = Border(bottom=Side(style="thin", color="EEEEEE"),
                                 left=Side(style="thin", color="EEEEEE"),
                                 right=Side(style="thin", color="EEEEEE"))
            if col_i == 5 and ml_tot:    c.fill = pct_fill(mp);  c.font = pct_font(mp)
            elif col_i == 8 and rl_tot:  c.fill = pct_fill(rp);  c.font = pct_font(rp)
            elif col_i == 11 and ou_tot: c.fill = pct_fill(op);  c.font = pct_font(op)
        ws2.row_dimensions[row_idx].height = 20

    for col, w in {"B":16,"C":8,"D":8,"E":10,"F":10,"G":8,"H":8,"I":8,"J":8,"K":10}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "B13"

    # ── Sheet 3: Charts & Trends ────────────────────────────────────
    _add_charts_sheet(wb, daily)

    wb.save(EXCEL_PATH)
    print(f"  ✅ Excel saved to: {EXCEL_PATH}\n")


def _add_charts_sheet(wb: Workbook, daily_db_rows: list) -> None:
    """
    Build Charts & Trends sheet from daily_picks data.
    daily_db_rows: [(pick_date, ml_tot, ml_w, rl_tot, rl_w, ou_tot, ou_w), ...]
    ordered DESC by date — we reverse to get chronological order.
    """
    ws = wb.create_sheet("Charts & Trends")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "27AE60"

    # ── Title ────────────────────────────────────────────────────────
    ws.merge_cells("A1:Z1")
    hdr(ws, "A1", "MLB MODEL — PERFORMANCE CHARTS & TRENDS", NAVY, WHITE, size=13)
    ws.row_dimensions[1].height = 28

    if not daily_db_rows:
        ws.merge_cells("A2:Z2")
        hdr(ws, "A2", "No evaluated picks yet — charts will populate after first results.", MID, WHITE, size=10)
        return

    # Reverse to chronological order
    daily_chron = list(reversed(daily_db_rows))

    # ── Build daily stats ─────────────────────────────────────────────
    records = []
    cum_ml_w=0; cum_ml_t=0
    cum_rl_w=0; cum_rl_t=0
    cum_ou_w=0; cum_ou_t=0

    for pick_date, ml_tot, ml_w, rl_tot, rl_w, ou_tot, ou_w in daily_chron:
        ml_pct = round(ml_w/ml_tot, 4) if ml_tot else None
        rl_pct = round(rl_w/rl_tot, 4) if rl_tot else None
        ou_pct = round(ou_w/ou_tot, 4) if ou_tot else None
        cum_ml_w += ml_w;  cum_ml_t += ml_tot
        cum_rl_w += rl_w;  cum_rl_t += rl_tot
        cum_ou_w += ou_w;  cum_ou_t += ou_tot
        records.append({
            "date":    str(pick_date.strftime("%m/%d") if hasattr(pick_date, "strftime") else str(pick_date)[:5]),
            "ml_pct":  ml_pct,
            "rl_pct":  rl_pct,
            "ou_pct":  ou_pct,
            "ml_cum":  round(cum_ml_w/cum_ml_t, 4) if cum_ml_t >= 30 else None,
            "rl_cum":  round(cum_rl_w/cum_rl_t, 4) if cum_rl_t >= 30 else None,
            "ou_cum":  round(cum_ou_w/cum_ou_t, 4) if cum_ou_t >= 20 else None,
        })

    df = pd.DataFrame(records)
    df["ml_roll7"] = pd.Series(df["ml_pct"]).rolling(7, min_periods=3).mean().round(4)
    df["rl_roll7"] = pd.Series(df["rl_pct"]).rolling(7, min_periods=3).mean().round(4)
    df["ou_roll7"] = pd.Series(df["ou_pct"]).rolling(7, min_periods=3).mean().round(4)
    n = len(df)

    # ── Season summary stats header ───────────────────────────────────
    ml_tot_s = sum(r[1] for r in daily_chron)
    ml_w_s   = sum(r[2] for r in daily_chron)
    rl_tot_s = sum(r[3] for r in daily_chron)
    rl_w_s   = sum(r[4] for r in daily_chron)
    ou_tot_s = sum(r[5] for r in daily_chron)
    ou_w_s   = sum(r[6] for r in daily_chron)

    ml_pct_s = ml_w_s/ml_tot_s if ml_tot_s else 0
    rl_pct_s = rl_w_s/rl_tot_s if rl_tot_s else 0
    ou_pct_s = ou_w_s/ou_tot_s if ou_tot_s else 0

    ws.merge_cells("A2:Z2")
    hdr(ws, "A2",
        f"Season 2026  |  ML: {ml_w_s}-{ml_tot_s-ml_w_s} ({ml_pct_s:.1%})  |  "
        f"RL: {rl_w_s}-{rl_tot_s-rl_w_s} ({rl_pct_s:.1%})  |  "
        f"O/U: {ou_w_s}-{ou_tot_s-ou_w_s} ({ou_pct_s:.1%})  |  "
        f"Break-even: 52.4%  |  Sharp: 57%+",
        MID, WHITE, bold=False, size=10)
    ws.row_dimensions[2].height = 20

    # ── Data table 1: Rolling (cols A-E) ─────────────────────────────
    roll_hdrs = ["Date","ML 7d","RL 7d","O/U 7d","B-Even"]
    for i, h in enumerate(roll_hdrs):
        c = ws.cell(4, i+1, h)
        c.font = Font(name="Arial", bold=True, size=8, color=WHITE)
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i+1)].width = 8

    for i, row in df.iterrows():
        r = i + 5
        ws.row_dimensions[r].height = 13
        ws.cell(r,1).value = row["date"]
        ws.cell(r,1).font  = Font(name="Arial", size=7)
        ws.cell(r,1).alignment = Alignment(horizontal="center")
        for j, col in enumerate(["ml_roll7","rl_roll7","ou_roll7"], start=2):
            c = ws.cell(r, j)
            c.value = row[col]
            c.number_format = "0%"
            c.font = Font(name="Arial", size=7)
            c.alignment = Alignment(horizontal="center")
        ws.cell(r,5).value = 0.524
        ws.cell(r,5).number_format = "0%"
        ws.cell(r,5).font = Font(name="Arial", size=7)

    # ── Data table 2: Cumulative (cols G-K) ──────────────────────────
    cum_hdrs = ["Date","ML Cum","RL Cum","O/U Cum","B-Even"]
    for i, h in enumerate(cum_hdrs):
        c = ws.cell(4, i+7, h)
        c.font = Font(name="Arial", bold=True, size=8, color=WHITE)
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i+7)].width = 8

    for i, row in df.iterrows():
        r = i + 5
        ws.cell(r,7).value = row["date"]
        ws.cell(r,7).font  = Font(name="Arial", size=7)
        ws.cell(r,7).alignment = Alignment(horizontal="center")
        for j, col in enumerate(["ml_cum","rl_cum","ou_cum"], start=1):
            c = ws.cell(r, 7+j)
            c.value = row[col]
            c.number_format = "0%"
            c.font = Font(name="Arial", size=7)
            c.alignment = Alignment(horizontal="center")
        ws.cell(r,11).value = 0.524
        ws.cell(r,11).number_format = "0%"
        ws.cell(r,11).font = Font(name="Arial", size=7)

    # ── Data table 3: Daily ML bars (cols M-N) ───────────────────────
    for h, col in [("M4","Date"),("N4","ML %")]:
        c = ws[h]; c.value = col
        c.font = Font(name="Arial", bold=True, size=8, color=WHITE)
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["M"].width = 8
    ws.column_dimensions["N"].width = 8

    for i, row in df.iterrows():
        r = i + 5
        ws.cell(r,13).value = row["date"]
        ws.cell(r,13).font  = Font(name="Arial", size=7)
        ws.cell(r,13).alignment = Alignment(horizontal="center")
        ws.cell(r,14).value = row["ml_pct"]
        ws.cell(r,14).number_format = "0%"
        ws.cell(r,14).font = Font(name="Arial", size=7)

    # ── CHART 1: 7-Day Rolling Win Rate ──────────────────────────────
    chart1 = LineChart()
    chart1.title = "7-Day Rolling Win Rate"
    chart1.style = 10
    chart1.y_axis.numFmt = "0%"
    chart1.y_axis.scaling.min = 0.25
    chart1.y_axis.scaling.max = 0.85
    chart1.y_axis.title = "Win Rate"
    chart1.legend.position = "b"
    chart1.width = 20; chart1.height = 11

    for col_idx, color, dash in [(2,"1B2A4A",None),(3,"27AE60",None),(4,"F0A500",None),(5,"E74C3C","dash")]:
        ref = Reference(ws, min_col=col_idx, min_row=4, max_row=4+n)
        chart1.add_data(ref, titles_from_data=True)
        s = chart1.series[-1]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = 18000 if not dash else 12000
        if dash: s.graphicalProperties.line.dashDot = dash
        s.smooth = True

    chart1.set_categories(Reference(ws, min_col=1, min_row=5, max_row=4+n))
    chart1.x_axis.tickLblSkip = max(1, n//8)
    ws.add_chart(chart1, "A52")

    # ── CHART 2: Cumulative Win Rate ──────────────────────────────────
    chart2 = LineChart()
    chart2.title = "Cumulative Win Rate — Season to Date"
    chart2.style = 10
    chart2.y_axis.numFmt = "0%"
    chart2.y_axis.scaling.min = 0.42
    chart2.y_axis.scaling.max = 0.70
    chart2.y_axis.title = "Win Rate"
    chart2.legend.position = "b"
    chart2.width = 20; chart2.height = 11

    for col_idx, color, dash in [(8,"1B2A4A",None),(9,"27AE60",None),(10,"F0A500",None),(11,"E74C3C","dash")]:
        ref = Reference(ws, min_col=col_idx, min_row=4, max_row=4+n)
        chart2.add_data(ref, titles_from_data=True)
        s = chart2.series[-1]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = 18000 if not dash else 12000
        if dash: s.graphicalProperties.line.dashDot = dash
        s.smooth = True

    chart2.set_categories(Reference(ws, min_col=7, min_row=5, max_row=4+n))
    chart2.x_axis.tickLblSkip = max(1, n//8)
    ws.add_chart(chart2, "L52")

    # ── CHART 3: Daily ML Win Rate bars ──────────────────────────────
    chart3 = BarChart()
    chart3.title = "Daily ML Win Rate"
    chart3.style = 10; chart3.type = "col"; chart3.grouping = "clustered"
    chart3.y_axis.numFmt = "0%"
    chart3.y_axis.scaling.min = 0; chart3.y_axis.scaling.max = 1.0
    chart3.legend.position = "b"
    chart3.width = 20; chart3.height = 11

    chart3.add_data(Reference(ws, min_col=14, min_row=4, max_row=4+n), titles_from_data=True)
    chart3.series[-1].graphicalProperties.solidFill = NAVY
    chart3.set_categories(Reference(ws, min_col=13, min_row=5, max_row=4+n))
    chart3.x_axis.tickLblSkip = max(1, n//8)
    ws.add_chart(chart3, "A70")

    # ── Key insights box ──────────────────────────────────────────────
    ws.merge_cells("L70:Z70")
    hdr(ws, "L70", "KEY INSIGHTS", MID, WHITE, size=10)
    ws.row_dimensions[70].height = 20

    be = 0.524; sharp = 0.570
    insights = [
        f"ML: {ml_w_s}-{ml_tot_s-ml_w_s} ({ml_pct_s:.1%}) — {'✅ Above' if ml_pct_s>=be else '❌ Below'} break-even",
        f"RL: {rl_w_s}-{rl_tot_s-rl_w_s} ({rl_pct_s:.1%}) — {'🎯 Sharp' if rl_pct_s>=sharp else ('✅ Above' if rl_pct_s>=be else '❌ Below')} target",
        f"O/U: {ou_w_s}-{ou_tot_s-ou_w_s} ({ou_pct_s:.1%}) — {'✅ Above' if ou_pct_s>=be else '❌ Below'} break-even",
        "Charts update automatically every morning",
        "Break-even: 52.4% | Sharp target: 57%+",
        "Lineup features gain weight at ~June 20 retrain",
    ]
    for i, note in enumerate(insights):
        r = 71 + i
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"L{r}:Z{r}")
        c = ws[f"L{r}"]
        c.value = note
        c.font  = Font(name="Arial", size=9,
                       italic=(i >= 3),
                       color=GREEN_FONT if "✅" in note or "🎯" in note else (RED_FONT if "❌" in note else "555555"))
        c.alignment = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"\n{'='*55}")
    print(f"  daily_picks_tracker.py  |  {date.today()}")
    print(f"{'='*55}\n")

    c = conn()
    try:
        c.autocommit = False
        with c.cursor() as cur:
            ensure_tables(cur)
            c.commit()

            evaluate_yesterday(cur)
            c.commit()

            save_today_picks(cur)
            c.commit()

            export_excel(cur)

        print("✅ Done.")

    except Exception:
        c.rollback()
        raise
    finally:
        c.close()


if __name__ == "__main__":
    main()
